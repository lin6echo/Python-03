import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active

sheet.merge_cells("A2:D4")

sheet.cell(row=2, column=1).value = "Twelve cells join together."

# merge cell C6 and D6

sheet.merge_cells('C6:D6')

sheet.cell(row=6, column=6).value = "Two merge cells"

wb.save("Excel/sample2.xlsx")