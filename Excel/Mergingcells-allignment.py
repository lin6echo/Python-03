from openpyxl.styles import Alignment
from openpyxl import Workbook
      
wb = Workbook()  
sheet = wb.active  
      
sheet.merge_cells('A1:B2')  
      
cell = sheet.cell(row=1, column=1)  
cell.value = 'Devansh Sharma'  
cell.alignment = Alignment(horizontal='center', vertical='center')  
      
wb.save('merging.xlsx')  