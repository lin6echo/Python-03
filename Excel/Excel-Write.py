# import openpyxl module
import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

c1 = sheet.cell(row=1, column=1)
# writing values to cells
c1.value = "Hello"

c2 = sheet.cell(row=1, column=2)
c2.value = "World"

c3 = sheet["A2"]
c3.value = "Welcome"

# B2 means column = 2 & row = 2.
c4 = sheet["B2"]
c4.value = "Everyone"

wb.save("Python-03/Excel/sample.xlsx")