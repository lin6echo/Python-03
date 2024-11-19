import openpyxl

wb = openpyxl.load_workbook("Excel/sample2.xlsx")
sheet = wb.active

# unmerge the cells
sheet.unmerge_cells("A2:D4")

sheet.unmerge_cells("C6:D6")

wb.save("Excel/sample3.xlsx")