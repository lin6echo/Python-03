import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

# writing to the specified cell
sheet.cell(row=1, column=1).value = "hello"
sheet.cell(row=2, column=2).value = "everyone"

# set the height of the row
sheet.row_dimensions[1].height = 70

# set the width of the column
sheet.column_dimensions["B"].width = 20

# save the file
wb.save("Python-03/Excel/sample1.xlsx")