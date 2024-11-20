import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import Workbook

wb = Workbook
ws = wb.active
ws['A1'].fill = PatternFill(bgColor="FFC7CE", fill_type = "solid")

wb.save("Excel/sample8.xlsx")